import openpyxl


class Train(object):
    def __init__(self):
        # 打开信息库
        self.f = openpyxl.load_workbook('列车信息库.xlsx')
        # 获取工作表
        self.sheet = self.f['Sheet1']

# 信息录入
    def info_record(self, li_info):
        # 获取信息库中的发车时间
        stime_1 = self.sheet['E'][1:]
        # 比较录入的列车发车时间并排序
        # 获取录入的发车时间
        stime_2 = li_info[4]
        # 通过for循环将信息库中的列车发车时间和录入的列车发车时间进行比较
        for i in range(len(stime_1)):
            # 如果录入发车的时间比信息库中的发车的时间早，则在其前面插入一行
            # a.__ge__(b) 判断a是否大于等于b
            if stime_2.__ge__(stime_1[i].value) == False:
                self.sheet.insert_rows(idx=i+2) # 如果发车的时间早，则在其前面插入一行，录入信息
                # 录入信息
                for j in range(len(li_info)):
                    self.sheet.cell(i+2, j+1).value = li_info[j]
                # 录入信息后退出
                break
            elif i == len(stime_1)-1: # 如果录入的发车时间比其他时间都晚，则在最后一行录入信息
                for j in range(len(li_info)):
                    self.sheet.cell(len(stime_1)+2, j+1).value = li_info[j]
                break
        # 信息保存
        self.f.save('列车信息库.xlsx')


# 查询所有信息
    def info_search_all(self):
        # 获取所有行的数据
        rows = self.sheet.rows
        # 定义一个空列表用来保存数据
        li_info = []
        for row in rows:
            # 将获取的每一行信息放入li_temp中，最后再添加到li_info中
            li_temp = [i.value for i in row]
            # 保存每一行数据
            li_info.append(li_temp)
        # 返回结果
        return li_info

# 根据某一信息查询（车次、起点站和终点站）
    def info_search(self, info):
        li_info = [] # 存放查找到的列车信息
        res = [] # 对查找到的列车信息去重，并保存到res中，作为返回结果
        # 获取信息库中的信息
        cols = self.sheet['A:C']
        if info[0] != '':
            # 根据车次号查询
            for cell_1 in cols[0]:
                if info[0] == cell_1.value: # 如果输入的车次号存在
                    # 获取信息所在的行数，以及信息，并保存到列表li中，
                    li_info.append([j.value for j in self.sheet[cell_1.row]])

        if info[1] != '' and info[2] != '':
            for col_1 in cols[1]:
                if info[1] == col_1.value:
                    for col_2 in cols[2]:
                        if col_2.value == info[2]:
                            li_info.append([i.value for i in self.sheet[col_2.row]])

        # 去重操作
        [res.append(i) for i in li_info if i not in res]
        return res


# 修改列车信息
    def search(self, info): # 先通过用户输入的车次号信息进行查找，并返回结果，如果存在则用户修改信息后运行info_modify函数，如果不存在则提示用户信息不存在
        li = []
        col = self.sheet['A']
        for i in col:
            if i.value == info:
                li = [j.value for j in self.sheet[i.row]]
        return li

    def info_modify(self,li_info): # li_info为用户修改后的信息，获取用户修改后的信息，然后根据此信息修改
        # 根据车次查找所在位置
        col = self.sheet['A']
        for i in col:
            if i.value == li_info[0]:
                row_info = self.sheet[i.row]
                for j in range(len(row_info)):
                    self.sheet.cell(i.row, j+1).value = li_info[j]
                break
        # 信息保存
        self.f.save('F:/pythonproject/作业/列车系统/列车信息库.xlsx')


# 删除信息
    def del_info(self, info):
        # 根据车次查找所在位置并删除
        col = self.sheet['A']
        for i in col:
            if i.value == info:
                self.sheet.delete_rows(i.row)












